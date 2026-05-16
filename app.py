from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
from hmac import compare_digest
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components

from core.data_loader import load_formula_catalog, load_materials, paper_options
from core.exporter import make_quote_dataframe
from core.formulas import calculate_edge_protector, calculate_tube
from core.models import EdgeProtectorInput, TubeInput


st.set_page_config(
    page_title="IP Poland Pricing Engine",
    page_icon="🧮",
    layout="wide",
    initial_sidebar_state="expanded",
)


def inject_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --ip-navy: #102A43;
            --ip-blue: #1D5F99;
            --ip-blue-soft: #E8F2FB;
            --ip-kraft: #B77838;
            --ip-green: #2F7D55;
            --ip-bg: #F5F7FA;
            --ip-card: #FFFFFF;
            --ip-border: #D9E2EC;
            --ip-text: #162033;
            --ip-muted: #64748B;
        }

        @keyframes fadeSlideUp {
            from {
                opacity: 0;
                transform: translateY(10px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        html, body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
            background: var(--ip-bg);
        }

        .stApp {
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            color: var(--ip-text);
        }

        .block-container {
            max-width: 1380px;
            padding-top: 1.4rem;
            padding-bottom: 3rem;
            animation: fadeSlideUp 0.35s ease-out;
        }

        /* Sidebar */
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #102A43 0%, #123B66 100%);
            border-right: 1px solid rgba(255, 255, 255, 0.12);
        }

        section[data-testid="stSidebar"] h1,
        section[data-testid="stSidebar"] h2,
        section[data-testid="stSidebar"] h3,
        section[data-testid="stSidebar"] p,
        section[data-testid="stSidebar"] span,
        section[data-testid="stSidebar"] label {
            color: #FFFFFF !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label {
            background: rgba(255, 255, 255, 0.08);
            border: 1px solid rgba(255, 255, 255, 0.14);
            border-radius: 12px;
            padding: 9px 11px;
            margin-bottom: 7px;
            transition: background 0.2s ease, transform 0.2s ease;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
            background: rgba(255, 255, 255, 0.18);
            transform: translateX(2px);
        }

        /* Hero */
        .hero {
            background: #FFFFFF;
            border: 1px solid var(--ip-border);
            border-radius: 20px;
            padding: 30px 34px;
            margin-bottom: 24px;
            box-shadow: 0 10px 30px rgba(16, 42, 67, 0.08);
            border-left: 8px solid var(--ip-blue);
            position: relative;
            overflow: hidden;
        }

        .hero::after {
            content: "";
            position: absolute;
            right: -80px;
            top: -80px;
            width: 240px;
            height: 240px;
            border-radius: 50%;
            background: rgba(183, 120, 56, 0.13);
        }

        .hero h1 {
            color: var(--ip-navy) !important;
            font-size: 40px;
            line-height: 1.08;
            font-weight: 800;
            letter-spacing: -0.7px;
            margin: 8px 0 10px 0;
        }

        .hero p {
            color: #526173 !important;
            font-size: 16px;
            line-height: 1.6;
            max-width: 1050px;
            margin: 0;
        }

        .badge {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 999px;
            background: var(--ip-blue-soft);
            color: var(--ip-blue);
            border: 1px solid #C9DCEB;
            font-weight: 800;
            font-size: 12px;
            margin-right: 8px;
            margin-bottom: 10px;
            letter-spacing: 0.25px;
        }

        .badge-kraft {
            background: #F7EBDD;
            color: #7A4A1D;
            border: 1px solid #E7C8A3;
        }

        h1, h2, h3, h4 {
            color: var(--ip-navy) !important;
            font-weight: 800 !important;
        }

        /* Cards */
        .glass-card {
            background: var(--ip-card);
            border: 1px solid var(--ip-border);
            border-radius: 18px;
            padding: 22px;
            box-shadow: 0 8px 24px rgba(16, 42, 67, 0.07);
        }

        /* Metric cards */
        div[data-testid="metric-container"] {
            background: #FFFFFF !important;
            border: 1px solid var(--ip-border) !important;
            border-radius: 16px !important;
            padding: 17px !important;
            box-shadow: 0 6px 18px rgba(16, 42, 67, 0.06) !important;
        }

        div[data-testid="metric-container"] label {
            color: #334155 !important;
            font-weight: 700 !important;
            opacity: 1 !important;
        }

        /* CRITICAL FIX: metric numbers */
        div[data-testid="metric-container"] [data-testid="stMetricValue"],
        div[data-testid="metric-container"] [data-testid="stMetricValue"] *,
        [data-testid="stMetricValue"],
        [data-testid="stMetricValue"] *,
        [data-testid="stMetricValue"] div,
        [data-testid="stMetricValue"] p,
        [data-testid="stMetricValue"] span {
            color: #102A43 !important;
            opacity: 1 !important;
            font-weight: 800 !important;
            text-shadow: none !important;
            filter: none !important;
        }

        div[data-testid="metric-container"] [data-testid="stMetricLabel"],
        div[data-testid="metric-container"] [data-testid="stMetricLabel"] *,
        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] * {
            color: #334155 !important;
            opacity: 1 !important;
            font-weight: 700 !important;
        }

        /* Inputs */
        div[data-testid="stTextInput"] input,
        div[data-testid="stNumberInput"] input,
        div[data-testid="stDateInput"] input {
            color: var(--ip-text) !important;
            background-color: #FFFFFF !important;
            border: 1px solid #C8D5E2 !important;
            border-radius: 10px !important;
            min-height: 42px !important;
        }

        /* Select boxes */
        div[data-baseweb="select"] {
            min-width: 100% !important;
        }

        div[data-baseweb="select"] > div {
            background-color: #FFFFFF !important;
            border: 1px solid #C8D5E2 !important;
            border-radius: 10px !important;
            min-height: 42px !important;
            color: var(--ip-text) !important;
        }

        div[data-baseweb="select"] span {
            color: var(--ip-text) !important;
            opacity: 1 !important;
        }

        div[data-baseweb="select"] svg {
            fill: var(--ip-navy) !important;
            color: var(--ip-navy) !important;
            opacity: 1 !important;
        }

        label {
            color: #334155 !important;
            font-weight: 700 !important;
        }

        /* Buttons */
        .stButton > button,
        .stDownloadButton > button {
            background: var(--ip-blue) !important;
            color: #FFFFFF !important;
            border: 1px solid var(--ip-blue) !important;
            border-radius: 11px !important;
            font-weight: 800 !important;
            min-height: 42px !important;
            box-shadow: 0 6px 16px rgba(29, 95, 153, 0.20);
            transition: transform 0.15s ease, background 0.15s ease;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            background: var(--ip-navy) !important;
            border-color: var(--ip-navy) !important;
            transform: translateY(-1px);
        }

        /* Dataframes */
        div[data-testid="stDataFrame"] {
            border: 1px solid var(--ip-border);
            border-radius: 14px;
            overflow: hidden;
            background: #FFFFFF;
        }

        .warning-box {
            padding: 16px 18px;
            border-radius: 14px;
            background: #FFF7E8;
            border: 1px solid #E7C27D;
            color: #5A3A0B;
        }

        .success-box {
            padding: 16px 18px;
            border-radius: 14px;
            background: #EDF9F1;
            border: 1px solid #A9D8B3;
            color: #14532D;
        }

        .footer {
            color: #64748B;
            font-size: 13px;
            padding-top: 30px;
            text-align: center;
        }

        hr {
            border-color: var(--ip-border);
        }
        /* Premium sidebar navigation buttons */
        section[data-testid="stSidebar"] div[role="radiogroup"] {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label {
            background: rgba(255, 255, 255, 0.08) !important;
            border: 1px solid rgba(255, 255, 255, 0.14) !important;
            border-radius: 14px !important;
            padding: 12px 14px !important;
            min-height: 48px !important;
            display: flex !important;
            align-items: center !important;
            transition: all 0.18s ease-in-out !important;
            cursor: pointer !important;
            box-shadow: 0 8px 18px rgba(0, 0, 0, 0.10);
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
            background: rgba(255, 255, 255, 0.18) !important;
            transform: translateX(3px);
            border-color: rgba(255, 255, 255, 0.30) !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label span {
            color: #FFFFFF !important;
            font-size: 16px !important;
            font-weight: 800 !important;
            letter-spacing: -0.2px;
        }

        /* Hide default radio circles */
        section[data-testid="stSidebar"] div[role="radiogroup"] input {
            display: none !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label > div:first-child {
            display: none !important;
        }

        /* Selected item highlight */
        section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
            background: linear-gradient(135deg, #27603B 0%, #3C7A4D 100%) !important;
            border-color: rgba(220, 240, 160, 0.45) !important;
            box-shadow: 0 10px 24px rgba(39, 96, 59, 0.30);
        }
        /* Sidebar polish: compact logo, cleaner nav, less scrolling */
        section[data-testid="stSidebar"] img {
            max-height: 78px !important;
            width: auto !important;
            object-fit: contain !important;
            margin: 0 auto 8px auto !important;
            display: block !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] {
            gap: 7px !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label {
            min-height: 44px !important;
            padding: 9px 12px !important;
            border-radius: 14px !important;
            background: rgba(255, 255, 255, 0.09) !important;
            border: 1px solid rgba(255, 255, 255, 0.14) !important;
            box-shadow: 0 6px 14px rgba(0, 0, 0, 0.10) !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
            background: rgba(255, 255, 255, 0.18) !important;
            border-color: rgba(255, 255, 255, 0.28) !important;
            transform: translateX(2px);
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label span {
            font-size: 15px !important;
            font-weight: 800 !important;
            color: #FFFFFF !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] input {
            display: none !important;
        }

        section[data-testid="stSidebar"] div[role="radiogroup"] label > div:first-child {
            display: none !important;
        }

        section[data-testid="stSidebar"] h1 {
            font-size: 24px !important;
            line-height: 1.15 !important;
            margin-bottom: 4px !important;
        }

        section[data-testid="stSidebar"] hr {
            margin-top: 18px !important;
            margin-bottom: 14px !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

def require_login() -> None:
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return

    st.set_page_config(
        page_title="Inter-Pack Pricing Engine Login",
        page_icon="🔐",
        layout="centered",
    )

    st.markdown(
        """
        <div style="
            max-width: 520px;
            margin: 70px auto 20px auto;
            padding: 34px;
            border-radius: 24px;
            background: linear-gradient(135deg, #17120D 0%, #241C15 100%);
            color: white;
            box-shadow: 0 24px 70px rgba(23,18,13,0.28);
            text-align: center;
        ">
            <img 
                src="https://inter-pack.com.pl/wp-content/uploads/2021/10/logo-lightpng.png"
                style="max-width: 260px; margin-bottom: 22px;"
            />
            <h1 style="margin: 0; font-size: 30px; color: white;">
                Pricing Engine
            </h1>
            <p style="color: rgba(255,255,255,0.70); margin-top: 10px;">
                Secure access for authorized users
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login", use_container_width=True)

    if submitted:
        expected_username = st.secrets.get("APP_USERNAME", "")
        expected_password = st.secrets.get("APP_PASSWORD", "")

        username_ok = compare_digest(username, expected_username)
        password_ok = compare_digest(password, expected_password)

        if username_ok and password_ok:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Invalid username or password.")

    st.stop()
    
def interpack_brand_header() -> None:
    components.html(
        """
        <div class="ip-shell">
            <style>
                :root {
                    --ip-black: #17120D;
                    --ip-dark: #241C15;
                    --ip-green: #27603B;
                    --ip-green-light: #DCF0A0;
                    --ip-card: rgba(255, 255, 255, 0.08);
                    --ip-border: rgba(255, 255, 255, 0.14);
                    --ip-text-soft: rgba(255, 255, 255, 0.68);
                    --ip-text-muted: rgba(255, 255, 255, 0.50);
                }

                @keyframes ipFadeUp {
                    0% {
                        opacity: 0;
                        transform: translateY(18px);
                    }
                    100% {
                        opacity: 1;
                        transform: translateY(0);
                    }
                }

                @keyframes ipFloatOne {
                    0% {
                        transform: translate3d(0, 0, 0) rotate(0deg);
                    }
                    50% {
                        transform: translate3d(-12px, 10px, 0) rotate(4deg);
                    }
                    100% {
                        transform: translate3d(0, 0, 0) rotate(0deg);
                    }
                }

                @keyframes ipFloatTwo {
                    0% {
                        transform: translate3d(0, 0, 0) scale(1);
                    }
                    50% {
                        transform: translate3d(12px, -12px, 0) scale(1.04);
                    }
                    100% {
                        transform: translate3d(0, 0, 0) scale(1);
                    }
                }

                @keyframes ipPulse {
                    0% {
                        box-shadow: 0 0 0 0 rgba(220, 240, 160, 0.18);
                    }
                    70% {
                        box-shadow: 0 0 0 14px rgba(220, 240, 160, 0);
                    }
                    100% {
                        box-shadow: 0 0 0 0 rgba(220, 240, 160, 0);
                    }
                }

                * {
                    box-sizing: border-box;
                }

                body {
                    margin: 0;
                    padding: 0;
                }

                .ip-shell {
                    width: 100%;
                    min-height: 560px;
                    position: relative;
                    overflow: hidden;
                    border-radius: 32px;
                    padding: 34px;
                    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
                    background:
                        radial-gradient(circle at 8% 10%, rgba(220, 240, 160, 0.14), transparent 26%),
                        radial-gradient(circle at 92% 10%, rgba(39, 96, 59, 0.34), transparent 30%),
                        radial-gradient(circle at 50% 120%, rgba(220, 240, 160, 0.10), transparent 36%),
                        linear-gradient(135deg, #17120D 0%, #241C15 46%, #0E1511 100%);
                    color: #FFFFFF;
                    border: 1px solid rgba(255, 255, 255, 0.12);
                    box-shadow: 0 26px 90px rgba(23, 18, 13, 0.34);
                    animation: ipFadeUp 0.5s ease-out;
                }

                .ip-orb-one {
                    position: absolute;
                    width: 420px;
                    height: 420px;
                    right: -180px;
                    top: -200px;
                    border-radius: 50%;
                    background: rgba(220, 240, 160, 0.13);
                    animation: ipFloatOne 8s ease-in-out infinite;
                }

                .ip-orb-two {
                    position: absolute;
                    width: 320px;
                    height: 320px;
                    left: -150px;
                    bottom: -150px;
                    border-radius: 50%;
                    background: rgba(39, 96, 59, 0.28);
                    animation: ipFloatTwo 9s ease-in-out infinite;
                }

                .ip-grid-lines {
                    position: absolute;
                    inset: 0;
                    opacity: 0.08;
                    background-image:
                        linear-gradient(rgba(255,255,255,0.18) 1px, transparent 1px),
                        linear-gradient(90deg, rgba(255,255,255,0.18) 1px, transparent 1px);
                    background-size: 44px 44px;
                    mask-image: linear-gradient(180deg, rgba(0,0,0,0.85), transparent);
                }

                .ip-content {
                    position: relative;
                    z-index: 3;
                }

                .ip-topbar {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 24px;
                    flex-wrap: wrap;
                    margin-bottom: 34px;
                }

                .ip-logo-zone {
                    display: flex;
                    align-items: center;
                    gap: 18px;
                    flex-wrap: wrap;
                }

                .ip-logo-card {
                    height: 72px;
                    min-width: 270px;
                    padding: 14px 20px;
                    border-radius: 20px;
                    background: rgba(255,255,255,0.07);
                    border: 1px solid rgba(255,255,255,0.15);
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    backdrop-filter: blur(10px);
                }

                .ip-logo-card img {
                    height: 42px;
                    max-width: 230px;
                    object-fit: contain;
                    display: block;
                }

                .ip-vertical-line {
                    width: 1px;
                    height: 54px;
                    background: rgba(255,255,255,0.18);
                }

                .ip-top-copy {
                    display: flex;
                    flex-direction: column;
                    gap: 4px;
                }

                .ip-kicker {
                    font-size: 12px;
                    line-height: 1.2;
                    text-transform: uppercase;
                    letter-spacing: 1.8px;
                    font-weight: 900;
                    color: rgba(255,255,255,0.58);
                }

                .ip-top-title {
                    font-size: 17px;
                    line-height: 1.25;
                    color: #FFFFFF;
                    font-weight: 900;
                }

                .ip-status-zone {
                    display: flex;
                    align-items: center;
                    gap: 10px;
                    flex-wrap: wrap;
                    justify-content: flex-end;
                }

                .ip-status-pill {
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    padding: 10px 13px;
                    border-radius: 999px;
                    background: rgba(255,255,255,0.08);
                    border: 1px solid rgba(255,255,255,0.14);
                    color: #FFFFFF;
                    font-size: 12px;
                    font-weight: 900;
                    white-space: nowrap;
                    backdrop-filter: blur(8px);
                }

                .ip-status-pill.green {
                    background: rgba(39,96,59,0.34);
                    border-color: rgba(111,190,132,0.35);
                    color: #E9FBEF;
                }

                .ip-status-dot {
                    width: 9px;
                    height: 9px;
                    border-radius: 50%;
                    background: #6BE58D;
                    animation: ipPulse 2.2s infinite;
                }

                .ip-main {
                    display: grid;
                    grid-template-columns: minmax(0, 1.35fr) minmax(340px, 0.8fr);
                    gap: 34px;
                    align-items: stretch;
                }

                .ip-hero-copy {
                    min-width: 0;
                    padding-top: 6px;
                }

                .ip-eyebrow {
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    color: var(--ip-green-light);
                    font-size: 13px;
                    line-height: 1.2;
                    font-weight: 900;
                    letter-spacing: 1.5px;
                    text-transform: uppercase;
                    margin-bottom: 14px;
                }

                .ip-eyebrow-mark {
                    width: 28px;
                    height: 2px;
                    background: var(--ip-green-light);
                    border-radius: 999px;
                }

                .ip-heading {
                    margin: 0 0 18px 0;
                    max-width: 900px;
                    color: #FFFFFF;
                    font-size: clamp(38px, 5vw, 68px);
                    line-height: 0.98;
                    font-weight: 950;
                    letter-spacing: -2px;
                }

                .ip-heading span {
                    color: var(--ip-green-light);
                }

                .ip-description {
                    max-width: 890px;
                    margin: 0;
                    color: var(--ip-text-soft);
                    font-size: 16px;
                    line-height: 1.78;
                    font-weight: 500;
                }

                .ip-product-cloud {
                    margin-top: 24px;
                    display: flex;
                    flex-wrap: wrap;
                    gap: 10px;
                }

                .ip-chip {
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    padding: 10px 12px;
                    border-radius: 14px;
                    background: rgba(255,255,255,0.08);
                    border: 1px solid rgba(255,255,255,0.12);
                    color: rgba(255,255,255,0.90);
                    font-size: 13px;
                    font-weight: 900;
                    white-space: nowrap;
                }

                .ip-chip.highlight {
                    background: rgba(220,240,160,0.13);
                    border-color: rgba(220,240,160,0.28);
                    color: #F3FFD2;
                }

                .ip-side-panel {
                    background: rgba(255,255,255,0.075);
                    border: 1px solid rgba(255,255,255,0.13);
                    border-radius: 26px;
                    padding: 18px;
                    backdrop-filter: blur(12px);
                    box-shadow: inset 0 1px 0 rgba(255,255,255,0.08);
                }

                .ip-panel-title {
                    color: #FFFFFF;
                    font-size: 16px;
                    line-height: 1.25;
                    font-weight: 950;
                    margin-bottom: 16px;
                }

                .ip-panel-subtitle {
                    color: rgba(255,255,255,0.56);
                    font-size: 12px;
                    line-height: 1.5;
                    margin-bottom: 18px;
                }

                .ip-feature-list {
                    display: grid;
                    gap: 12px;
                }

                .ip-feature {
                    display: grid;
                    grid-template-columns: 40px 1fr;
                    gap: 12px;
                    align-items: flex-start;
                    padding: 12px;
                    border-radius: 16px;
                    background: rgba(255,255,255,0.08);
                    border: 1px solid rgba(255,255,255,0.10);
                }

                .ip-feature-icon {
                    width: 40px;
                    height: 40px;
                    border-radius: 13px;
                    background: rgba(220,240,160,0.13);
                    border: 1px solid rgba(220,240,160,0.22);
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-size: 18px;
                }

                .ip-feature-title {
                    color: #FFFFFF;
                    font-size: 13px;
                    line-height: 1.25;
                    font-weight: 950;
                    margin-bottom: 4px;
                }

                .ip-feature-text {
                    color: rgba(255,255,255,0.62);
                    font-size: 12px;
                    line-height: 1.5;
                    font-weight: 500;
                }

                .ip-bottom {
                    margin-top: 30px;
                    padding-top: 20px;
                    border-top: 1px solid rgba(255,255,255,0.12);
                    display: grid;
                    grid-template-columns: 1fr auto;
                    gap: 18px;
                    align-items: center;
                }

                .ip-bottom-note {
                    color: var(--ip-text-muted);
                    font-size: 12px;
                    line-height: 1.6;
                    font-weight: 600;
                }

                .ip-bottom-actions {
                    display: flex;
                    gap: 8px;
                    flex-wrap: wrap;
                    justify-content: flex-end;
                }

                .ip-mini-badge {
                    padding: 8px 11px;
                    border-radius: 999px;
                    background: rgba(255,255,255,0.08);
                    border: 1px solid rgba(255,255,255,0.12);
                    color: rgba(255,255,255,0.76);
                    font-size: 12px;
                    font-weight: 900;
                    white-space: nowrap;
                }

                @media (max-width: 960px) {
                    .ip-shell {
                        padding: 24px;
                        min-height: auto;
                    }

                    .ip-main {
                        grid-template-columns: 1fr;
                    }

                    .ip-vertical-line {
                        display: none;
                    }

                    .ip-heading {
                        font-size: 40px;
                    }

                    .ip-bottom {
                        grid-template-columns: 1fr;
                    }

                    .ip-bottom-actions {
                        justify-content: flex-start;
                    }

                    .ip-status-zone {
                        justify-content: flex-start;
                    }
                }
            </style>

            <div class="ip-orb-one"></div>
            <div class="ip-orb-two"></div>
            <div class="ip-grid-lines"></div>

            <div class="ip-content">
                <div class="ip-topbar">
                    <div class="ip-logo-zone">
                        <div class="ip-logo-card">
                            <img src="https://inter-pack.com.pl/wp-content/uploads/2021/10/logo-lightpng.png" alt="Inter-Pack logo">
                        </div>

                        <div class="ip-vertical-line"></div>

                        <div class="ip-top-copy">
                            <div class="ip-kicker">Internal pricing cockpit</div>
                            <div class="ip-top-title">Production-grade calculator for packaging quotations</div>
                        </div>
                    </div>

                    <div class="ip-status-zone">
                        <span class="ip-status-pill">📦 Packaging modules</span>
                        <span class="ip-status-pill">📊 Quote analytics</span>
                        <span class="ip-status-pill green"><span class="ip-status-dot"></span> Engine online</span>
                    </div>
                </div>

                <div class="ip-main">
                    <div class="ip-hero-copy">
                        <div class="ip-eyebrow">
                            <span class="ip-eyebrow-mark"></span>
                            Inter‑Pack Polonia · Packaging Intelligence
                        </div>

                        <h1 class="ip-heading">
                            Smart pricing for <span>custom cardboard packaging</span>
                        </h1>

                        <p class="ip-description">
                            A modern calculation workspace for edge protectors, cardboard tubes,
                            cores, material parameters, palletization, validation, quotation exports,
                            and controlled Excel-to-Python formula migration.
                        </p>

                        <div class="ip-product-cloud">
                            <span class="ip-chip highlight">🧱 Edge protectors</span>
                            <span class="ip-chip">🌀 Tubes & cores</span>
                            <span class="ip-chip">📄 Paper & cardboard</span>
                            <span class="ip-chip">📦 Cardboard pallet</span>
                            <span class="ip-chip">🏭 Production</span>
                            <span class="ip-chip">✅ Validation</span>
                        </div>
                    </div>

                    <div class="ip-side-panel">
                        <div class="ip-panel-title">Business-ready calculation suite</div>
                        <div class="ip-panel-subtitle">
                            Built for pricing precision, production planning, and quotation workflows.
                        </div>

                        <div class="ip-feature-list">
                            <div class="ip-feature">
                                <div class="ip-feature-icon">⚙️</div>
                                <div>
                                    <div class="ip-feature-title">Formula engine</div>
                                    <div class="ip-feature-text">
                                        Python pricing logic connected to Excel migration benchmarks.
                                    </div>
                                </div>
                            </div>

                            <div class="ip-feature">
                                <div class="ip-feature-icon">🧮</div>
                                <div>
                                    <div class="ip-feature-title">Product calculators</div>
                                    <div class="ip-feature-text">
                                        Edge protector and tube/core calculators with commercial outputs.
                                    </div>
                                </div>
                            </div>

                            <div class="ip-feature">
                                <div class="ip-feature-icon">📤</div>
                                <div>
                                    <div class="ip-feature-title">Export-ready quotes</div>
                                    <div class="ip-feature-text">
                                        CSV and styled Excel exports for customer and internal workflows.
                                    </div>
                                </div>
                            </div>

                            <div class="ip-feature">
                                <div class="ip-feature-icon">🧬</div>
                                <div>
                                    <div class="ip-feature-title">Formula audit</div>
                                    <div class="ip-feature-text">
                                        Workbook formula distribution and validation roadmap.
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>

                <div class="ip-bottom">
                    <div class="ip-bottom-note">
                        Inspired by Inter‑Pack’s packaging product lines: cores, edge protectors,
                        tubes, paper and cardboard, cardboard pallets, and production workflows.
                    </div>

                    <div class="ip-bottom-actions">
                        <span class="ip-mini-badge">FSC-ready</span>
                        <span class="ip-mini-badge">ISO quality</span>
                        <span class="ip-mini-badge">EU delivery</span>
                    </div>
                </div>
            </div>
        </div>
        """,
        height=820,
        scrolling=False,
    )
    
def hero() -> None:
    st.markdown(
        """
        <div class="hero">
            <span class="badge">STREAMLIT PRICING ENGINE</span>
            <span class="badge">IP POLAND 2025</span>
            <h1>Industrial Packaging Calculator</h1>
            <p>
                Professional pricing cockpit for Edge Protectors, Tubes/Cores,
                palletization, strength estimation, margin simulation, material data,
                and Excel-to-Python formula migration.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def money(value: float) -> str:
    return f"{value:,.4f} PLN"

def convert_from_pln(
    value_pln: float,
    currency: str,
) -> float:
    fx_rates = load_materials().get("fx_rates", {})

    if currency == "PLN":
        return value_pln

    rate = fx_rates.get(currency)

    if not rate:
        return value_pln

    return value_pln / rate


def money_in_currency(
    value_pln: float,
    currency: str,
) -> str:
    converted_value = convert_from_pln(
        value_pln=value_pln,
        currency=currency,
    )

    return f"{converted_value:,.4f} {currency}"


def number(value: float, digits: int = 3) -> str:
    return f"{value:,.{digits}f}"


def download_quote(product: str, inputs: dict, result: dict) -> None:
    quote_df = make_quote_dataframe(
        product=product,
        inputs=inputs,
        result=result,
    )

    csv = quote_df.to_csv(index=False).encode("utf-8")

    st.download_button(
        label="⬇️ Download quote CSV",
        data=csv,
        file_name=f"{product.lower().replace(' ', '_')}_quote.csv",
        mime="text/csv",
        use_container_width=True,
    )

def create_quote_excel_file(
    product: str,
    inputs: dict,
    result: dict,
) -> bytes:
    quote_df = make_quote_dataframe(
        product=product,
        inputs=inputs,
        result=result,
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        quote_df.to_excel(
            writer,
            index=False,
            sheet_name="Quote",
            startrow=4,
        )

        worksheet = writer.sheets["Quote"]

        # Title area
        worksheet.merge_cells("A1:C1")
        worksheet["A1"] = "IP Poland Pricing Engine"
        worksheet["A1"].font = Font(
            color="FFFFFF",
            bold=True,
            size=16,
        )
        worksheet["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor="0F172A",
        )
        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.merge_cells("A2:C2")
        worksheet["A2"] = "Professional Quote Export"
        worksheet["A2"].font = Font(
            color="0F172A",
            bold=True,
            size=12,
        )
        worksheet["A2"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.merge_cells("A3:C3")
        worksheet["A3"] = f"Product: {product}"
        worksheet["A3"].font = Font(
            color="334155",
            italic=True,
        )
        worksheet["A3"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # Table header styling
        header_row = 5

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="0F172A",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Body styling
        body_alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = body_alignment

        # Column widths
        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 32
        worksheet.column_dimensions["C"].width = 42

        # Row heights
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[2].height = 22
        worksheet.row_dimensions[3].height = 22
        worksheet.row_dimensions[header_row].height = 22

        # Freeze panes and filter
        worksheet.freeze_panes = "A6"
        worksheet.auto_filter.ref = f"A5:C{worksheet.max_row}"

    output.seek(0)

    return output.getvalue()
    
def result_chart(result: dict, product: str) -> None:
    chart_values = [
        result.get("price_per_rm_pln", 0),
        result.get("price_per_piece_pln", 0),
        result.get("price_per_kg_pln", 0),
    ]

    chart_labels = [
        "Price / r.m.",
        "Price / piece",
        "Price / kg",
    ]

    chart_colors = [
        "#1F6FB2",
        "#C99A5B",
        "#2F855A",
    ]

    fig = go.Figure()

    fig.add_trace(
        go.Bar(
            x=chart_labels,
            y=chart_values,
            marker=dict(
                color=chart_colors,
                line=dict(
                    color="#FFFFFF",
                    width=1.5,
                ),
            ),
            text=[
                f"{value:,.4f}"
                for value in chart_values
            ],
            textposition="outside",
            textfont=dict(
                color="#172033",
                size=14,
                family="Arial",
            ),
            hovertemplate="<b>%{x}</b><br>Value: %{y:.4f} PLN<extra></extra>",
        )
    )

    fig.update_layout(
        title=dict(
            text=f"{product} price composition",
            font=dict(
                size=20,
                color="#12335B",
                family="Arial",
            ),
            x=0.02,
            xanchor="left",
        ),
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#FFFFFF",
        font=dict(
            color="#172033",
            family="Arial",
            size=13,
        ),
        height=430,
        margin=dict(
            l=40,
            r=30,
            t=70,
            b=60,
        ),
        bargap=0.35,
        showlegend=False,
        yaxis=dict(
            title=dict(
                text="Value in PLN",
                font=dict(
                    color="#5E6B7A",
                    size=13,
                ),
            ),
            gridcolor="#E3EAF2",
            zerolinecolor="#C9D6E3",
            tickfont=dict(
                color="#344256",
                size=12,
            ),
        ),
        xaxis=dict(
            tickfont=dict(
                color="#344256",
                size=13,
            ),
            title=None,
        ),
    )

    fig.update_traces(
        cliponaxis=False,
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
        config={
            "displayModeBar": False,
            "responsive": True,
        },
    )


def validation_page() -> None:
    st.subheader("✅ Validation Center")

    st.markdown(
        """
        This page compares the Streamlit calculation results against benchmark
        values extracted from the original Excel calculator.
        """
    )

    st.markdown("### Edge Protector benchmark")

    edge_input = EdgeProtectorInput(
        side_1_mm=25.0,
        side_2_mm=150.0,
        thickness_mm=1.8,
        length_mm=500.0,
        product_type="IPP-30",
        outer_cover="Testliner Grey",
        quantity_pcs=33000,
        quantity_per_pallet=1000,
        transport_cost_pln=0.0,
        margin_percent=0.0,
    )

    edge_result = calculate_edge_protector(edge_input)

    edge_benchmarks = [
        {
            "Metric": "Weight kg/r.m.",
            "Excel Benchmark": 0.235290752,
            "App Result": edge_result.weight_kg_per_rm,
        },
        {
            "Metric": "Price per r.m.",
            "Excel Benchmark": 0.05778908933942856,
            "App Result": edge_result.price_per_rm_pln,
        },
        {
            "Metric": "Price per piece",
            "Excel Benchmark": 0.02889454466971428,
            "App Result": edge_result.price_per_piece_pln,
        },
        {
            "Metric": "Three-point bending N",
            "Excel Benchmark": 301.5770911471299,
            "App Result": edge_result.three_point_bending_n,
        },
    ]

    edge_df = pd.DataFrame(edge_benchmarks)

    edge_df["Difference"] = (
        edge_df["App Result"] - edge_df["Excel Benchmark"]
    )

    edge_df["Difference %"] = (
        edge_df["Difference"] / edge_df["Excel Benchmark"] * 100
    )

    edge_df["Status"] = edge_df["Difference %"].abs().apply(
        lambda value: "PASS" if value <= 0.5 else "REVIEW"
    )

    st.dataframe(
        edge_df,
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("### Tube / Core benchmark")

    tube_input = TubeInput(
        diameter_mm=76.8,
        thickness_mm=5.8,
        length_mm=525.0,
        product_type="IPP-30",
        inner_cover="nie",
        outer_cover="Testliner Grey",
        quantity_pcs=27324,
        quantity_per_pallet=37,
        transport_cost_pln=0.0,
        margin_percent=0.0,
    )

    tube_result = calculate_tube(tube_input)

    tube_benchmarks = [
        {
            "Metric": "Weight kg/r.m.",
            "Excel Benchmark": 1.1144761904761904,
            "App Result": tube_result.weight_kg_per_rm,
        },
        {
            "Metric": "Price per r.m.",
            "Excel Benchmark": 0.2380301683113101,
            "App Result": tube_result.price_per_rm_pln,
        },
        {
            "Metric": "Price per piece",
            "Excel Benchmark": 0.12496583836343782,
            "App Result": tube_result.price_per_piece_pln,
        },
        {
            "Metric": "Flat crush N/0.1m",
            "Excel Benchmark": 655.236437504587,
            "App Result": tube_result.flat_crush_n_per_0_1m,
        },
    ]

    tube_df = pd.DataFrame(tube_benchmarks)

    tube_df["Difference"] = (
        tube_df["App Result"] - tube_df["Excel Benchmark"]
    )

    tube_df["Difference %"] = (
        tube_df["Difference"] / tube_df["Excel Benchmark"] * 100
    )

    tube_df["Status"] = tube_df["Difference %"].abs().apply(
        lambda value: "PASS" if value <= 0.5 else "REVIEW"
    )

    st.dataframe(
        tube_df,
        use_container_width=True,
        hide_index=True,
    )

    st.markdown(
        """
        <div class='warning-box'>
            PASS means the app result is within ±0.5% of the Excel benchmark.
            REVIEW means the formula should be checked or recalibrated.
        </div>
        """,
        unsafe_allow_html=True,
    )

def quote_builder_page() -> None:
    st.subheader("🧾 Quote Builder")

    st.markdown(
        """
        Build a quick customer quote from the available calculator modules.
        This page will be expanded step by step into a full quotation tool.
        """
    )

    st.markdown("### Customer and commercial information")

    customer_col, sales_col, payment_col = st.columns(3)

    customer_name = customer_col.text_input(
        "Customer name",
        value="",
        placeholder="Enter customer name",
    )

    salesperson = sales_col.text_input(
        "Salesperson",
        value="",
        placeholder="Enter salesperson",
    )

    payment_days = payment_col.number_input(
        "Payment terms, days",
        min_value=0,
        value=10,
        step=1,
    )

    quote_meta_1, quote_meta_2, quote_meta_3 = st.columns(3)

    quote_id = quote_meta_1.text_input(
        "Quote ID",
        value="Q-2025-001",
        placeholder="Example: Q-2025-001",
    )

    quote_date = quote_meta_2.date_input(
        "Quote date",
    )

    selected_currency = quote_meta_3.selectbox(
        "Display currency",
        ["PLN", "EUR", "USD", "UAH"],
        index=0,
    )

    quote_date_text = quote_date.strftime("%Y-%m-%d")

    st.divider()

    st.markdown("### Product selection")

    product = st.selectbox(
        "Select product type",
        [
            "Edge Protector",
            "Tube / Core",
        ],
    )

    if product == "Edge Protector":
        st.markdown("### Edge Protector quote input")

        papers = paper_options()

        c1, c2, c3 = st.columns(3)

        product_type = c1.selectbox(
            "Type",
            papers,
            index=papers.index("IPP-30") if "IPP-30" in papers else 0,
            key="quote_edge_type",
        )

        outer_cover = c2.selectbox(
            "Outer cover",
            papers,
            index=papers.index("Testliner Grey") if "Testliner Grey" in papers else 0,
            key="quote_edge_outer",
        )

        quantity = c3.number_input(
            "Quantity, pcs",
            min_value=1,
            value=33000,
            step=100,
            key="quote_edge_qty",
        )

        d1, d2, d3, d4 = st.columns(4)

        side_1 = d1.number_input(
            "Side 1, mm",
            min_value=1.0,
            value=25.0,
            step=1.0,
            key="quote_edge_side_1",
        )

        side_2 = d2.number_input(
            "Side 2, mm",
            min_value=1.0,
            value=150.0,
            step=1.0,
            key="quote_edge_side_2",
        )

        thickness = d3.number_input(
            "Thickness, mm",
            min_value=0.1,
            value=1.8,
            step=0.1,
            key="quote_edge_thickness",
        )

        length = d4.number_input(
            "Length, mm",
            min_value=1.0,
            value=500.0,
            step=10.0,
            key="quote_edge_length",
        )

        commercial_1, commercial_2, commercial_3 = st.columns(3)

        qty_per_pallet = commercial_1.number_input(
            "Qty per pallet",
            min_value=1,
            value=1000,
            step=10,
            key="quote_edge_qty_per_pallet",
        )

        transport = commercial_2.number_input(
            "Transport, PLN",
            min_value=0.0,
            value=0.0,
            step=50.0,
            key="quote_edge_transport",
        )

        margin = commercial_3.number_input(
            "Margin, %",
            min_value=-100.0,
            value=0.0,
            step=1.0,
            key="quote_edge_margin",
        )

        quote_input = EdgeProtectorInput(
            side_1_mm=side_1,
            side_2_mm=side_2,
            thickness_mm=thickness,
            length_mm=length,
            product_type=product_type,
            outer_cover=outer_cover,
            quantity_pcs=int(quantity),
            quantity_per_pallet=int(qty_per_pallet),
            transport_cost_pln=transport,
            margin_percent=margin,
        )

        quote_result = calculate_edge_protector(quote_input)
        quote_result_dict = quote_result.to_dict()

        st.markdown("### Edge Protector quote result")

        r1, r2, r3, r4 = st.columns(4)

        r1.metric(
            "Price / r.m.",
            money_in_currency(
                quote_result.price_per_rm_pln,
                selected_currency,
            ),
        )

        r2.metric(
            "Price / piece",
            money_in_currency(
                quote_result.price_per_piece_pln,
                selected_currency,
            ),
        )

        r3.metric(
            "Total value",
            money_in_currency(
                quote_result.total_value_pln,
                selected_currency,
            ),
        )

        r4.metric(
            "Net weight",
            f"{quote_result.net_weight_kg:,.2f} kg",
        )

        result_chart(
            quote_result_dict,
            "Edge Protector Quote",
        )

    else:
        st.markdown("### Tube / Core quote input")

        papers = paper_options()

        c1, c2, c3 = st.columns(3)

        product_type = c1.selectbox(
            "Type",
            papers,
            index=papers.index("IPP-30") if "IPP-30" in papers else 0,
            key="quote_tube_type",
        )

        outer_cover = c2.selectbox(
            "Outer cover",
            papers,
            index=papers.index("Testliner Grey") if "Testliner Grey" in papers else 0,
            key="quote_tube_outer",
        )

        inner_cover = c3.selectbox(
            "Inner cover",
            ["nie"] + papers,
            index=0,
            key="quote_tube_inner",
        )

        d1, d2, d3, d4 = st.columns(4)

        diameter = d1.number_input(
            "Diameter, mm",
            min_value=1.0,
            value=76.8,
            step=0.1,
            key="quote_tube_diameter",
        )

        thickness = d2.number_input(
            "Thickness, mm",
            min_value=0.1,
            value=5.8,
            step=0.1,
            key="quote_tube_thickness",
        )

        length = d3.number_input(
            "Length, mm",
            min_value=1.0,
            value=525.0,
            step=5.0,
            key="quote_tube_length",
        )

        quantity = d4.number_input(
            "Quantity, pcs",
            min_value=1,
            value=27324,
            step=100,
            key="quote_tube_qty",
        )

        commercial_1, commercial_2, commercial_3 = st.columns(3)

        qty_per_pallet = commercial_1.number_input(
            "Qty per pallet",
            min_value=1,
            value=37,
            step=1,
            key="quote_tube_qty_per_pallet",
        )

        transport = commercial_2.number_input(
            "Transport, PLN",
            min_value=0.0,
            value=0.0,
            step=50.0,
            key="quote_tube_transport",
        )

        margin = commercial_3.number_input(
            "Margin, %",
            min_value=-100.0,
            value=0.0,
            step=1.0,
            key="quote_tube_margin",
        )

        quote_input = TubeInput(
            diameter_mm=diameter,
            thickness_mm=thickness,
            length_mm=length,
            product_type=product_type,
            inner_cover=inner_cover,
            outer_cover=outer_cover,
            quantity_pcs=int(quantity),
            quantity_per_pallet=int(qty_per_pallet),
            transport_cost_pln=transport,
            margin_percent=margin,
        )

        quote_result = calculate_tube(quote_input)
        quote_result_dict = quote_result.to_dict()

        st.markdown("### Tube / Core quote result")

        r1, r2, r3, r4 = st.columns(4)

        r1.metric(
            "Price / r.m.",
            money_in_currency(
                quote_result.price_per_rm_pln,
                selected_currency,
            ),
        )

        r2.metric(
            "Price / piece",
            money_in_currency(
                quote_result.price_per_piece_pln,
                selected_currency,
            ),
        )

        r3.metric(
            "Total value",
            money_in_currency(
                quote_result.total_value_pln,
                selected_currency,
            ),
        )

        r4.metric(
            "Net weight",
            f"{quote_result.net_weight_kg:,.2f} kg",
        )

        result_chart(
            quote_result_dict,
            "Tube / Core Quote",
        )

    st.divider()

    st.markdown("### Quote summary")

    summary_df = pd.DataFrame(
        [
            {
                "Field": "Quote ID",
                "Value": quote_id,
            },
            {
                "Field": "Quote date",
                "Value": quote_date_text,
            },
            {
                "Field": "Customer",
                "Value": customer_name,
            },
            {
                "Field": "Salesperson",
                "Value": salesperson,
            },
            {
                "Field": "Payment terms",
                "Value": f"{payment_days} days",
            },
            {
                "Field": "Selected product",
                "Value": product,
            },
            {
                "Field": "Quantity",
                "Value": quote_input.quantity_pcs,
            },
            {
                "Field": "Price per running meter, PLN",
                "Value": quote_result_dict["price_per_rm_pln"],
            },
            {
                "Field": "Price per piece, PLN",
                "Value": quote_result_dict["price_per_piece_pln"],
            },
            {
                "Field": "Total value, PLN",
                "Value": quote_result_dict["total_value_pln"],
            },
            {
                "Field": "Net weight, kg",
                "Value": quote_result_dict["net_weight_kg"],
            },
            {
                "Field": "Pallets",
                "Value": quote_result_dict["pallets"],
            },
        ]
    )

    st.dataframe(
        summary_df,
        use_container_width=True,
        hide_index=True,
    )

    export_df = make_quote_dataframe(
        product=product,
        inputs={        
            "quote_id": quote_id,
            "quote_date": quote_date_text,
            "customer_name": customer_name,
            "salesperson": salesperson,
            "payment_days": payment_days,
            **asdict(quote_input),
        },
        result=quote_result_dict,
    )

    csv = export_df.to_csv(index=False).encode("utf-8")

    st.download_button(
        label="⬇️ Download quote CSV",
        data=csv,
        file_name="quote_builder_export.csv",
        mime="text/csv",
        use_container_width=True,
    )
    excel_bytes = create_quote_excel_file(
        product=product,
        inputs={
            "quote_id": quote_id,
            "quote_date": quote_date_text,
            "customer_name": customer_name,
            "salesperson": salesperson,
            "payment_days": payment_days,
            **asdict(quote_input),
        },
        result=quote_result_dict,
    )

    st.download_button(
        label="⬇️ Download quote Excel",
        data=excel_bytes,
        file_name="quote_builder_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    preview_df = pd.DataFrame(
        [
            {
                "Field": "Customer",
                "Value": customer_name,
            },
            {
                "Field": "Salesperson",
                "Value": salesperson,
            },
            {
                "Field": "Payment terms",
                "Value": f"{payment_days} days",
            },
            {
                "Field": "Selected product",
                "Value": product,
            },
        ]
    )

    st.markdown("### Quote setup preview")

    st.dataframe(
        preview_df,
        use_container_width=True,
        hide_index=True,
    )
def dashboard_page() -> None:
    st.subheader("🚀 Command Center")

    catalog = load_formula_catalog()

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("Formula cells loaded", f"{catalog.get('total_formulas', 0):,}")
    c2.metric("Product modules", "4+")
    c3.metric("Current MVP", "Edge + Tubes")
    c4.metric("Currency", "PLN")

    st.markdown("### Migration status")

    status = pd.DataFrame(
    [
        {
            "Module": "Edge Protectors",
            "Status": "MVP formula engine ready",
            "Priority": "High",
        },
        {
            "Module": "Tubes / Cores",
            "Status": "MVP formula engine ready",
            "Priority": "High",
        },
        {
            "Module": "Paper database",
            "Status": "JSON database ready",
            "Priority": "High",
        },
        {
            "Module": "Formula Audit",
            "Status": "Workbook formula distribution loaded",
            "Priority": "High",
        },
        {
            "Module": "Validation Center",
            "Status": "Excel benchmark comparison ready",
            "Priority": "High",
        },
        {
            "Module": "Cardboard Pallets",
            "Status": "Next implementation",
            "Priority": "Medium",
        },
        {
            "Module": "HoneyComb",
            "Status": "High-complexity module planned later",
            "Priority": "Medium",
        },
    ]
)

    st.dataframe(status, use_container_width=True, hide_index=True)


def edge_page() -> None:
    st.subheader("🧱 Edge Protector Calculator")

    left, right = st.columns([0.38, 0.62], gap="large")

    with left:
        st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
        st.markdown("#### Product inputs")

        papers = paper_options()

        product_type = st.selectbox(
            "Type",
            papers,
            index=papers.index("IPP-30") if "IPP-30" in papers else 0,
        )

        outer_cover = st.selectbox(
            "Outer cover",
            papers,
            index=papers.index("Testliner Grey") if "Testliner Grey" in papers else 0,
        )

        c1, c2 = st.columns(2)

        side_1 = c1.number_input(
            "Side 1 / x1, mm",
            min_value=1.0,
            value=25.0,
            step=1.0,
        )

        side_2 = c2.number_input(
            "Side 2 / x2, mm",
            min_value=1.0,
            value=150.0,
            step=1.0,
        )

        c3, c4 = st.columns(2)

        thickness = c3.number_input(
            "Wall thickness, mm",
            min_value=0.1,
            value=1.8,
            step=0.1,
        )

        length = c4.number_input(
            "Length, mm",
            min_value=1.0,
            value=500.0,
            step=10.0,
        )

        c5, c6 = st.columns(2)

        qty = c5.number_input(
            "Quantity, pcs",
            min_value=1,
            value=33000,
            step=100,
        )

        qty_per_pallet = c6.number_input(
            "Qty per pallet",
            min_value=1,
            value=1000,
            step=10,
        )

        st.markdown("#### Commercial")

        c7, c8 = st.columns(2)

        transport = c7.number_input(
            "Transport, PLN",
            min_value=0.0,
            value=0.0,
            step=50.0,
        )

        margin = c8.number_input(
            "Margin, %",
            min_value=-100.0,
            value=0.0,
            step=1.0,
        )

        st.markdown("</div>", unsafe_allow_html=True)

    inp = EdgeProtectorInput(
        side_1_mm=side_1,
        side_2_mm=side_2,
        thickness_mm=thickness,
        length_mm=length,
        product_type=product_type,
        outer_cover=outer_cover,
        quantity_pcs=int(qty),
        quantity_per_pallet=int(qty_per_pallet),
        transport_cost_pln=transport,
        margin_percent=margin,
    )

    result = calculate_edge_protector(inp)
    result_dict = result.to_dict()

    with right:
        m1, m2, m3, m4 = st.columns(4)

        m1.metric("Price / r.m.", money(result.price_per_rm_pln))
        m2.metric("Price / piece", money(result.price_per_piece_pln))
        m3.metric("Weight kg/r.m.", number(result.weight_kg_per_rm, 6))
        m4.metric("Bending force", f"{result.three_point_bending_n:,.1f} N")

        result_chart(result_dict, "Edge Protector")

        a, b = st.columns(2)

        with a:
            st.markdown("#### Logistics")

            logistics = pd.DataFrame(
                [
                    {"Metric": "Total running meters", "Value": result.total_running_m},
                    {"Metric": "Net weight kg", "Value": result.net_weight_kg},
                    {"Metric": "Pallets", "Value": result.pallets},
                    {"Metric": "Pallet weight kg", "Value": result.pallet_weight_kg},
                ]
            )

            st.dataframe(logistics, use_container_width=True, hide_index=True)

        with b:
            st.markdown("#### Production order")

            st.json(
                {
                    "product": "Kątownik / Edge Protector",
                    "dimensions": f"{side_1} x {side_2} x {thickness} / {length} mm",
                    "type": product_type,
                    "outer_cover": outer_cover,
                    "quantity_pcs": int(qty),
                }
            )

        download_quote(
            product="Edge Protector",
            inputs=asdict(inp),
            result=result_dict,
        )


def tube_page() -> None:
    st.subheader("🌀 Tubes / Cores Calculator")

    left, right = st.columns([0.38, 0.62], gap="large")

    with left:
        st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
        st.markdown("#### Product inputs")

        papers = paper_options()

        product_type = st.selectbox(
            "Type",
            papers,
            index=papers.index("IPP-30") if "IPP-30" in papers else 0,
            key="tube_type",
        )

        inner_cover = st.selectbox(
            "Inner cover",
            ["nie"] + papers,
            index=0,
        )

        outer_cover = st.selectbox(
            "Outer cover",
            papers,
            index=papers.index("Testliner Grey") if "Testliner Grey" in papers else 0,
            key="tube_outer",
        )

        c1, c2 = st.columns(2)

        diameter = c1.number_input(
            "Diameter, mm",
            min_value=1.0,
            value=76.8,
            step=0.1,
        )

        thickness = c2.number_input(
            "Wall thickness, mm",
            min_value=0.1,
            value=5.8,
            step=0.1,
            key="tube_thickness",
        )

        c3, c4 = st.columns(2)

        length = c3.number_input(
            "Length, mm",
            min_value=1.0,
            value=525.0,
            step=5.0,
            key="tube_length",
        )

        qty_per_pallet = c4.number_input(
            "Qty per pallet",
            min_value=1,
            value=37,
            step=1,
            key="tube_qty_per_pallet",
        )

        c5, c6 = st.columns(2)

        qty = c5.number_input(
            "Quantity, pcs",
            min_value=1,
            value=27324,
            step=100,
            key="tube_qty",
        )

        transport = c6.number_input(
            "Transport, PLN",
            min_value=0.0,
            value=0.0,
            step=50.0,
            key="tube_transport",
        )

        margin = st.number_input(
            "Margin, %",
            min_value=-100.0,
            value=0.0,
            step=1.0,
            key="tube_margin",
        )

        st.markdown("</div>", unsafe_allow_html=True)

    inp = TubeInput(
        diameter_mm=diameter,
        thickness_mm=thickness,
        length_mm=length,
        product_type=product_type,
        inner_cover=inner_cover,
        outer_cover=outer_cover,
        quantity_pcs=int(qty),
        quantity_per_pallet=int(qty_per_pallet),
        transport_cost_pln=transport,
        margin_percent=margin,
    )

    result = calculate_tube(inp)
    result_dict = result.to_dict()

    with right:
        m1, m2, m3, m4 = st.columns(4)

        m1.metric("Price / r.m.", money(result.price_per_rm_pln))
        m2.metric("Price / piece", money(result.price_per_piece_pln))
        m3.metric("Weight kg/r.m.", number(result.weight_kg_per_rm, 6))
        m4.metric("Flat crush", f"{result.flat_crush_n_per_0_1m:,.1f} N")

        result_chart(result_dict, "Tube / Core")

        st.markdown("#### Calculation summary")

        st.dataframe(
            pd.DataFrame(result_dict.items(), columns=["Metric", "Value"]),
            use_container_width=True,
            hide_index=True,
        )

        download_quote(
            product="Tube Core",
            inputs=asdict(inp),
            result=result_dict,
        )


def materials_page() -> None:
    st.subheader("📦 Materials & Calibration Database")

    data = load_materials()

    st.markdown(
        """
        <div class='warning-box'>
            This MVP stores material data in <b>data/materials.json</b>.
            Edit it in GitHub or add an admin save function later.
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("### Paper parameters")

    material_df = (
        pd.DataFrame(data["paper_types"])
        .T
        .reset_index()
        .rename(columns={"index": "material"})
    )

    st.dataframe(material_df, use_container_width=True, hide_index=True)

    st.markdown("### FX rates")
    st.json(data["fx_rates"])

    st.markdown("### Calibration constants")
    st.json(data["calibration"])


def formula_audit_page() -> None:
    st.subheader("🧬 Excel Formula Audit")

    catalog = load_formula_catalog()

    total_formulas = catalog.get("total_formulas", 0)
    sheet_stats = catalog.get("sheet_stats", [])
    formulas = catalog.get("formulas", [])

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("Extracted formula cells", f"{total_formulas:,}")
    c2.metric("Workbook sheets", f"{len(sheet_stats):,}")
    c3.metric(
        "Migrated MVP modules",
        "2",
        help="Edge Protectors and Tubes / Cores are currently included in the MVP engine.",
    )
    c4.metric(
        "Pending modules",
        "4+",
        help="HoneyComb, Cardboard Pallets, Technology sheets, and full formula validation.",
    )

    if not sheet_stats:
        st.info(
            "No formula catalog found yet. "
            "Create data/formula_catalog.json to display workbook formula statistics."
        )
        return

    st.markdown("### Workbook formula distribution")

    stats_df = pd.DataFrame(sheet_stats)

    expected_columns = ["sheet", "formulas", "rows", "cols", "status"]

    for col in expected_columns:
        if col not in stats_df.columns:
            stats_df[col] = ""

    st.dataframe(
        stats_df[expected_columns],
        use_container_width=True,
        hide_index=True,
    )

    chart_df = stats_df.copy()
    chart_df["formulas"] = pd.to_numeric(
        chart_df["formulas"],
        errors="coerce",
    ).fillna(0)

    fig = go.Figure()

    fig.add_trace(
        go.Bar(
            x=chart_df["sheet"],
            y=chart_df["formulas"],
            marker=dict(
                color="#1F6FB2",
                line=dict(
                    color="#FFFFFF",
                    width=1.5,
                ),
            ),
            text=[
                f"{value:,.0f}"
                for value in chart_df["formulas"]
            ],
            textposition="outside",
            textfont=dict(
                color="#172033",
                size=13,
                family="Arial",
            ),
            hovertemplate="<b>%{x}</b><br>Formula cells: %{y:,.0f}<extra></extra>",
        )
    )

    fig.update_layout(
        title=dict(
            text="Formula count by workbook sheet",
            font=dict(
                size=20,
                color="#12335B",
                family="Arial",
            ),
            x=0.02,
            xanchor="left",
        ),
        xaxis_title=None,
        yaxis_title="Formula cells",
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#FFFFFF",
        font=dict(
            color="#172033",
            family="Arial",
            size=13,
        ),
        height=460,
        margin=dict(
            l=50,
            r=30,
            t=70,
            b=130,
        ),
        bargap=0.32,
        showlegend=False,
        yaxis=dict(
            title=dict(
                text="Formula cells",
                font=dict(
                    color="#5E6B7A",
                    size=13,
                ),
            ),
            gridcolor="#E3EAF2",
            zerolinecolor="#C9D6E3",
            tickfont=dict(
                color="#344256",
                size=12,
            ),
        ),
        xaxis=dict(
            tickangle=-35,
            tickfont=dict(
                color="#344256",
                size=12,
            ),
        ),
    )

    fig.update_traces(
        cliponaxis=False,
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
        config={
            "displayModeBar": False,
            "responsive": True,
        },
    )

    st.markdown("### Migration interpretation")

    st.markdown(
        """
        <div class='warning-box'>
            The formula catalog is used as the Excel-to-Python migration map.
            It should not be pasted directly into the Streamlit UI or into one huge Python file.
            Instead, formulas should be migrated module by module into clean Python functions
            inside <b>core/formulas.py</b>.
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("### Migration priority")

    priority_df = pd.DataFrame(
        [
            {
                "Priority": 1,
                "Module": "Edge Protectors",
                "Reason": "Already included in the MVP formula engine.",
                "Next Action": "Validate against Excel benchmarks.",
            },
            {
                "Priority": 2,
                "Module": "Tubes / Cores",
                "Reason": "Already included in the MVP formula engine.",
                "Next Action": "Validate against Excel benchmarks.",
            },
            {
                "Priority": 3,
                "Module": "Paper Database",
                "Reason": "Material parameters are required by all calculators.",
                "Next Action": "Add admin editing and real price fields.",
            },
            {
                "Priority": 4,
                "Module": "Cardboard Pallets",
                "Reason": "Medium formula complexity and useful business module.",
                "Next Action": "Create pallet calculator page.",
            },
            {
                "Priority": 5,
                "Module": "HoneyComb",
                "Reason": "Highest formula complexity.",
                "Next Action": "Migrate after simpler modules are validated.",
            },
        ]
    )

    st.dataframe(
        priority_df,
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("### Formula search")

    if formulas:
        sheet_names = [item.get("sheet", "") for item in sheet_stats]

        selected_sheet = st.selectbox(
            "Filter by sheet",
            ["All"] + sheet_names,
        )

        filtered_formulas = formulas

        if selected_sheet != "All":
            filtered_formulas = [
                item
                for item in filtered_formulas
                if item.get("sheet") == selected_sheet
            ]

        query = st.text_input("Search formula text or cell reference")

        if query:
            q = query.lower()
            filtered_formulas = [
                item
                for item in filtered_formulas
                if q in str(item.get("formula", "")).lower()
                or q in str(item.get("cell", "")).lower()
            ]

        st.dataframe(
            pd.DataFrame(filtered_formulas[:1000]),
            use_container_width=True,
            hide_index=True,
        )

        st.caption("Showing first 1000 filtered formulas for performance.")
    else:
        st.info(
            "Lightweight formula catalog is loaded. "
            "Full individual formula search will become available after uploading "
            "the full extracted formula catalog."
        )


def main() -> None:
    require_login()
    inject_css()
    interpack_brand_header()
    hero()

    with st.sidebar:
        st.image(
            "https://inter-pack.com.pl/wp-content/uploads/2021/10/logo-lightpng.png",
            use_container_width=True,
        )

        if st.button("🔓 Logout", use_container_width=True):
            st.session_state.authenticated = False
            st.rerun()

        st.markdown(
            """
            <div style="padding: 10px 0 18px 0;">
                <div style="
                    color: white;
                    font-size: 26px;
                    font-weight: 900;
                    line-height: 1.15;
                    margin-top: 8px;
                ">
                    Pricing Engine
                </div>
                <div style="
                    color: rgba(255,255,255,0.70);
                    font-size: 13px;
                    font-weight: 600;
                    margin-top: 6px;
                    line-height: 1.4;
                ">
                    Industrial packaging calculator
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        page_labels = {
            "🏠 Dashboard": "Dashboard",
            "🧱 Edge Protector": "Edge Protector",
            "🌀 Tubes / Cores": "Tubes / Cores",
            "🧾 Quote Builder": "Quote Builder",
            "📦 Materials": "Materials",
            "🧬 Formula Audit": "Formula Audit",
            "✅ Validation": "Validation",
        }

        selected_page_label = st.radio(
            "Navigation",
            list(page_labels.keys()),
            label_visibility="collapsed",
        )

        page = page_labels[selected_page_label]

        st.divider()

        st.markdown(
            """
            <div style="
                color: rgba(255,255,255,0.65);
                font-size: 12px;
                font-weight: 700;
                line-height: 1.5;
                margin-top: 16px;
            ">
                <div style="margin-bottom: 8px;">Deployment</div>
                <div style="
                    color: white;
                    font-size: 13px;
                    font-weight: 800;
                    margin-bottom: 6px;
                ">
                    🟢 Streamlit Cloud
                </div>
                <div>
                    Built for GitHub deployment<br/>
                    Version 0.1.0 MVP
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    if page == "Dashboard":
        dashboard_page()
    elif page == "Edge Protector":
        edge_page()
    elif page == "Tubes / Cores":
        tube_page()
    elif page == "Quote Builder":
        quote_builder_page()
    elif page == "Materials":
        materials_page()
    elif page == "Formula Audit":
        formula_audit_page()
    elif page == "Validation":
        validation_page()
    else:
        st.error(f"Unknown page selected: {page}")

    st.markdown(
        """
        <div class='footer'>
            © IP Poland Pricing Engine · Excel-to-Python migration scaffold ·
            Validate all pricing before commercial use.
        </div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
